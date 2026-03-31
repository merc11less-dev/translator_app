import os
import sys
from typing import Dict, List, Optional
from difflib import get_close_matches


class DictionaryModel():
    def __init__(self, file_path: str = 'sources/words.xlsx'):
        self._dictionary: Dict[str, List[str]] = {}

        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
            full_path = os.path.join(base_path, file_path)
        else:
            base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            full_path = os.path.join(base_path, file_path)

        self._load_from_excel(full_path)

    def _load_from_excel(self, file_path: str) -> None:
        try:
            from openpyxl import load_workbook

            if not os.path.exists(file_path):
                print(f"Dictionary file not found: {file_path}")
                self._dictionary = {}
                return

            wb = load_workbook(file_path)
            sheet = wb.active

            self._dictionary = {}
            last_english_word = None

            for row_num in range(2, sheet.max_row + 1):
                english_cell = sheet.cell(row=row_num, column=1)
                russian_cell = sheet.cell(row=row_num, column=2)

                english_value = english_cell.value
                russian_value = russian_cell.value

                if english_value is None or str(english_value).strip() == "":
                    english_str = last_english_word
                else:
                    english_str = str(english_value).strip().lower()
                    last_english_word = english_str

                if russian_value is None or str(russian_value).strip() == "":
                    continue

                russian_str = str(russian_value).strip()
                if english_str and russian_str:
                    if english_str not in self._dictionary:
                        self._dictionary[english_str] = []
                    self._dictionary[english_str].append(russian_str)

            wb.close()
            print(f"Loaded {len(self._dictionary)} words from {file_path}")

        except Exception as e:
            print(f"Error loading dictionary: {e}")
            self._dictionary = {}

    def get_translation(self, word: str) -> Optional[str]:
        if not word or not isinstance(word, str):
            return None

        cleaned_word = word.strip().lower()
        if not cleaned_word:
            return None

        translations = self._dictionary.get(cleaned_word)
        if translations:
            return ', '.join(translations)

        return None

    def find_similar_words(self, word: str, cutoff: float = 0.6, limit: int = 10) -> List[str]:
        if not word or not isinstance(word, str):
            return []

        cleaned_word = word.strip().lower()
        if not cleaned_word:
            return []

        all_words = list(self._dictionary.keys())

        matches = get_close_matches(
            cleaned_word,
            all_words,
            n=limit,
            cutoff=cutoff
        )

        return matches

    def find_similar_words_with_priority(self, word: str, limit: int = 7) -> List[str]:
        if not word or not isinstance(word, str):
            return []

        cleaned_word = word.strip().lower()
        if not cleaned_word:
            return []

        all_words = list(self._dictionary.keys())

        exact_prefix = []
        first_word_prefix = []
        contains = []

        for w in all_words:
            w_lower = w.lower()

            if w_lower.startswith(cleaned_word):
                exact_prefix.append(w)
            elif ' ' in w_lower:
                first_word = w_lower.split()[0]
                if first_word.startswith(cleaned_word):
                    first_word_prefix.append(w)
            elif cleaned_word in w_lower:
                contains.append(w)

        close = get_close_matches(cleaned_word, all_words, n=limit, cutoff=0.5)

        result = []
        seen = set()

        for w in exact_prefix + first_word_prefix + contains + close:
            if w not in seen and len(result) < limit:
                result.append(w)
                seen.add(w)

        return result

    def get_all_words(self) -> List[str]:
        return list(self._dictionary.keys())

    def get_dictionary_size(self) -> int:
        return len(self._dictionary)